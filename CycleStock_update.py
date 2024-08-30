import numpy as np
import pandas as pd
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# READING FILES #
DanhSachSanPham = pd.read_excel('C:\\Users\\User\\Downloads\\Exercise\\Demand and Inventory for Special Stores\\26.08.2024\\SQL\\20240826_DanhSachSanPham.xlsx',sheet_name='DanhSachSP-NCC')
DanhSachNhomSanPham = pd.read_excel('C:\\Users\\User\\Downloads\\Exercise\\Demand and Inventory for Special Stores\\26.08.2024\\SQL\\20240826_DanhSachNhomSanPham.xlsx', sheet_name='Danh sách sản phẩm')
DiemTrangThai = pd.read_excel('C:\\Users\\User\\Downloads\\Exercise\\Demand and Inventory for Special Stores\\26.08.2024\\SQL\\Điểm trạng thái.xlsx')
ListSPNHCap5 = pd.read_excel('C:\\Users\\User\\Downloads\\Exercise\\Demand and Inventory for Special Stores\\26.08.2024\\SQL\\List SP NH cấp 5 Đồ chơi.xlsx')
DSNganhHangTinhThuong = pd.read_excel('C:\\Users\\User\\Downloads\\Exercise\\Demand and Inventory for Special Stores\\26.08.2024\\SQL\\20240826_DSNganhHangTinhThuong.xlsx')
DSThoiGianGiaoHangDenCuaHang = pd.read_excel('C:\\Users\\User\\Downloads\\Exercise\\Demand and Inventory for Special Stores\\26.08.2024\\SQL\\20240826_DSThoiGianGiaoHangDenCuaHang.xlsx')
InventoryRatio = pd.read_excel('C:\\Users\\User\\Downloads\\Exercise\\Demand and Inventory for Special Stores\\26.08.2024\\SQL\\20240826_InventoryRatio.xlsx')
SaleQuantityByStoreExcludedDefectID_7D = pd.read_excel('C:\\Users\\User\\Downloads\\Exercise\\Demand and Inventory for Special Stores\\26.08.2024\\SQL\\20240826_SaleQuantityByStoreExcludedDefectID_7D.xlsx')
SaleQuantityByStoreExcludedDefectID_30D = pd.read_excel('C:\\Users\\User\\Downloads\\Exercise\\Demand and Inventory for Special Stores\\26.08.2024\\SQL\\20240826_SaleQuantityByStoreExcludedDefectID_30D.xlsx')

# CREATE A NEW FILE FOR REPORT #
folder_path = 'C:\\Users\\User\\Downloads\\Exercise\\Demand and Inventory for Special Stores\\26.08.2024\\SQL'
data = {}
RawData = pd.DataFrame(data)
file_name = 'Nhu cầu của 6 stores đặc biệt.xlsx'
file_path = os.path.join(folder_path, file_name)
RawData.to_excel(file_path, index=False, engine='openpyxl')
wb = Workbook()
wb.remove(wb.active)
wb.create_sheet(title='CommonCAT')
wb.create_sheet(title='ToyBook')
wb.create_sheet(title='FashionAccessories')
wb.save(file_path)


# DATA PROCESSING #

    # Processing file Group ID #
        # Vlookup điểm trạng thái into DanhSachNhomSanPham #
lookup_dict = pd.Series(DiemTrangThai['Thứ tự ưu tiên'].values, index=DiemTrangThai['Trạng thái sản phẩm']).to_dict()
def vlookup(id_value):
    return lookup_dict.get(id_value)
DanhSachNhomSanPham['Điểm trạng thái sản phẩm'] = DanhSachNhomSanPham['Trạng thái sản phẩm'].apply(vlookup)
        # Sort Điểm trạng thái and Hệ số tỉ lệ % #
DanhSachNhomSanPham = DanhSachNhomSanPham.sort_values(by=['Điểm trạng thái sản phẩm', 'Hệ số tỉ lệ(%)'], ascending=[True, False])

    # Processing list SKUs #
        # Remove duplicate SKUs #
DanhSachSanPham = DanhSachSanPham.drop_duplicates(subset='Mã sản phẩm', keep='first')
        # Format Mã sản phẩm column at Text #
DanhSachSanPham['Mã sản phẩm'] = DanhSachSanPham['Mã sản phẩm'].astype(str).str.zfill(13)
        # Delete unrelated data #
DanhSachSanPham = DanhSachSanPham.drop(columns=['Mã tham chiếu', 'Số SP/thùng CC', 'Số SP/thùng NCC','Nhà cung cấp'])
        # Delete row if "Ngành hàng cấp 2" have GroupID in name #
DanhSachSanPham = DanhSachSanPham[~DanhSachSanPham['Ngành hàng cấp 2'].str.contains('GroupID', na=False)]
        # Filter CAT 1 #
DanhSachSanPham_CommonCAT = DanhSachSanPham[~DanhSachSanPham['Ngành hàng cấp 1'].str.contains(['Đồ chơi & Sách','Thời trang', 'Phụ kiện'], na=False)]
DanhSachSanPham_FashionAccessories = DanhSachSanPham[DanhSachSanPham['Ngành hàng cấp 1'].isin(['Thời trang', 'Phụ kiện'])]
DanhSachSanPham_ToyBook = DanhSachSanPham[DanhSachSanPham['Ngành hàng cấp 1'].isin('Đồ chơi & Sách')]
        
        # Copy All CAT to new file #
wb = openpyxl.load_workbook(file_path)
CommonCAT = wb['CommonCAT']
ToyBook = wb['ToyBook']
FashionAccessories = wb['FashionAccessories']
def write_listSKUsbyCAT_to_sheet(listSKUsbyCAT, sheet):
    for r_idx, row in enumerate(dataframe_to_rows(listSKUsbyCAT, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            # Set the column format to text if it is the 'Mã sản phẩm' column
            if listSKUsbyCAT.columns[c_idx-1] == 'Mã sản phẩm':
                cell.number_format = '@'  # Format as text

write_listSKUsbyCAT_to_sheet(DanhSachSanPham_CommonCAT, CommonCAT)
write_listSKUsbyCAT_to_sheet(DanhSachSanPham_FashionAccessories, FashionAccessories)
write_listSKUsbyCAT_to_sheet(DanhSachSanPham_ToyBook, ToyBook)

wb.save(file_path)

